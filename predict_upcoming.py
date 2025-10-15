# predict_upcoming.py
# Enhanced user interface for QB passing yards predictions with Excel output and embedded chart.
# Supports batch import from CSV and individual manual entry

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import BarChart, Reference, Series
import os

try:
    from src.predictor import predict_prop
    from src.config import PROPS_CONFIG
except ImportError as e:
    print(f"Error importing src modules: {e}. Ensure src/predictor.py and src/config.py exist.")
    exit(1)

# Valid NFL team codes
NFL_TEAMS = {'ARI', 'ATL', 'BAL', 'BUF', 'CAR', 'CHI', 'CIN', 'CLE', 'DAL', 'DEN', 'DET', 'GB', 'HOU', 'IND', 'JAX', 'KC', 'LA', 'LAC', 'LV', 'MIA', 'MIN', 'NE', 'NO', 'NYG', 'NYJ', 'PHI', 'PIT', 'SEA', 'SF', 'TB', 'TEN', 'WAS'}

def validate_team(team):
    """Validate team code."""
    if not team:
        raise ValueError("Team code cannot be empty.")
    team_upper = team.upper()
    if team_upper not in NFL_TEAMS:
        raise ValueError(f"Invalid team code: {team}. Must be one of {sorted(NFL_TEAMS)}")
    return team_upper

def validate_player(player_name, prop_name, player_data, season=None, prediction_week=None, min_games=10):
    """
    Validate player name exists and has sufficient historical data.
    
    Args:
        player_name: Player name to validate
        prop_name: Prop type being predicted
        player_data: DataFrame with historical player data
        season: Current season for prediction (optional, for stricter validation)
        prediction_week: Week being predicted (optional, for stricter validation)
        min_games: Minimum number of historical games required
    
    Returns:
        str: Validated player name
    
    Raises:
        ValueError: If player not found or insufficient data
    """
    if not player_name:
        raise ValueError("Player name cannot be empty.")
    
    # Check player exists
    if player_name not in player_data['player_name'].unique():
        raise ValueError(f"Player {player_name} not found in historical data.")
    
    # If season/week provided, do strict validation on available history
    if season is not None and prediction_week is not None:
        player_history = player_data[
            (player_data['player_name'] == player_name) & 
            ((player_data['season'] < season) | 
             ((player_data['season'] == season) & (player_data['week'] < prediction_week)))
        ]
    else:
        # Fallback: just check total games across all time
        player_history = player_data[player_data['player_name'] == player_name]
    
    game_count = len(player_history)
    
    if game_count < min_games:
        raise ValueError(
            f"Insufficient data for {player_name}: only {game_count} games found. "
            f"Minimum required: {min_games} games.\n"
            f"Model predictions are unreliable for players with limited NFL history.\n"
            f"Please select a player with more established data."
        )
    
    return player_name

def get_float_input(prompt):
    """Get valid float input."""
    while True:
        try:
            value = input(prompt).strip()
            if not value:
                raise ValueError("Input cannot be empty.")
            return float(value)
        except ValueError:
            print("Please enter a valid number.")

def get_int_input(prompt, min_val, max_val):
    """Get valid int input within range."""
    while True:
        try:
            value = input(prompt).strip()
            if not value:
                raise ValueError("Input cannot be empty.")
            val = int(value)
            if min_val <= val <= max_val:
                return val
            raise ValueError(f"Value must be between {min_val} and {max_val}.")
        except ValueError as e:
            print(f"Invalid input: {e}")

def get_existing_predictions(output_file, sheet_name):
    """Get existing predictions for the week to avoid duplicates."""
    if not os.path.exists(output_file):
        return set()
    
    try:
        existing_df = pd.read_excel(output_file, sheet_name=sheet_name)
        if 'player_name' in existing_df.columns:
            return set(existing_df['player_name'].dropna().unique())
    except (ValueError, KeyError, Exception):
        pass
    
    return set()

def create_excel_chart(ws, df, start_row, num_players):
    """Create embedded chart directly in Excel using openpyxl."""
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = f"Week {df['week'].iloc[0]} QB Passing Yards: Predictions vs Vegas Lines"
    chart.y_axis.title = 'Passing Yards'
    chart.x_axis.title = 'Player'
    chart.width = 20
    chart.height = 12

    # Data for predictions (column 5)
    pred_data = Reference(ws, min_col=5, min_row=start_row, max_row=start_row + num_players - 1)
    pred_series = Series(pred_data, title="Mean Prediction")
    chart.series.append(pred_series)

    # Data for Vegas lines (column 6)
    line_data = Reference(ws, min_col=6, min_row=start_row, max_row=start_row + num_players - 1)
    line_series = Series(line_data, title="Vegas Line")
    chart.series.append(line_series)

    # Player names as categories
    categories = Reference(ws, min_col=1, min_row=start_row, max_row=start_row + num_players - 1)
    chart.set_categories(categories)

    # Position chart below the data
    chart_cell = f"A{start_row + num_players + 2}"
    ws.add_chart(chart, chart_cell)
    
    return ws

def save_to_excel(df, week):
    """Save predictions to formatted Excel file with embedded chart."""
    output_file = os.path.join('data', 'passing-prop-predictions-2025.xlsx')
    sheet_name = f'Week {week}'
    
    df['week'] = week
    
    existing_players = get_existing_predictions(output_file, sheet_name)
    new_players_mask = ~df['player_name'].isin(existing_players)
    new_df = df[new_players_mask]
    
    if len(new_df) == 0:
        print(f"All {len(df)} predictions already exist in Week {week}. No new data added.")
        return
    
    if existing_players:
        print(f"Found {len(existing_players)} existing predictions for Week {week}.")
        print(f"Adding {len(new_df)} new predictions...")
    
    excel_cols = ['player_name', 'posteam', 'defteam', 'home_away', 'mean_prediction', 'line', 'actual', 'prob_over_line', 'lean', 'win_loss', 'p10', 'p50', 'p90']
    excel_df = new_df[excel_cols].copy()
    
    if os.path.exists(output_file):
        wb = openpyxl.load_workbook(output_file)
    else:
        wb = openpyxl.Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
    
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        existing_df = pd.read_excel(output_file, sheet_name=sheet_name)
        start_row = len(existing_df) + 2
        
        for row_idx, row in excel_df.iterrows():
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=start_row + row_idx, column=col_idx, value=value)
    else:
        ws = wb.create_sheet(sheet_name)
        start_row = 2
        
        headers = ['Player', 'Team', 'Opponent', 'Home/Away', 'Mean.Pred', 'Vegas.Line', 'Actual', 'Prob.Over', 'Lean', 'Win.Loss', 'P10', 'P50', 'P90']
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        for row_idx, row in excel_df.iterrows():
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=start_row + row_idx, column=col_idx, value=value)
    
    format_excel_sheet(ws, len(excel_df), start_row)
    
    if not existing_players:
        create_excel_chart(ws, df, start_row, len(excel_df))
    
    wb.save(output_file)
    print(f"Saved {len(new_df)} predictions to {output_file}, sheet '{sheet_name}'")
    if existing_players:
        print(f"Total predictions in Week {week}: {len(existing_players) + len(new_df)}")

def format_excel_sheet(ws, num_new_players, start_row):
    """Format the Excel sheet with styles and conditional formatting."""
    if start_row == 2 or ws.cell(row=1, column=1).value is None:
        headers = ['Player', 'Team', 'Opponent', 'Home/Away', 'Mean Pred', 'Vegas Line', 'Actual', 'Prob Over', 'Lean', 'W/L', 'P10', 'P50', 'P90']
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx).value = header
    
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    red_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
    header_font = Font(bold=True)
    
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    end_row = start_row + num_new_players - 1
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(1, 14):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal='center')
        
        # Format Lean column (I, column 9)
        lean_cell = ws.cell(row=row_idx, column=9)
        if lean_cell.value == 'OVER':
            lean_cell.fill = green_fill
        elif lean_cell.value == 'UNDER':
            lean_cell.fill = red_fill
        
        # Format W/L column (J, column 10)
        wl_cell = ws.cell(row=row_idx, column=10)
        if wl_cell.value == 'W':
            wl_cell.fill = green_fill
        elif wl_cell.value == 'L':
            wl_cell.fill = red_fill
    
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_length + 2

def load_batch_csv(csv_path):
    """Load predictions from CSV file.
    Expected columns: player_name, posteam, defteam, line, is_home
    """
    try:
        df = pd.read_csv(csv_path)
        required_cols = ['player_name', 'posteam', 'defteam', 'line', 'is_home']
        missing = [col for col in required_cols if col not in df.columns]
        if missing:
            raise ValueError(f"CSV missing required columns: {missing}")
        return df
    except Exception as e:
        raise ValueError(f"Error loading CSV: {e}")

def batch_predict(batch_df, prop_name, season, prediction_week, player_data):
    """Generate predictions for batch of players."""
    results = []
    
    # Get min_games from config
    config = PROPS_CONFIG[prop_name]
    min_games = config.get('min_games', 10)
    
    for idx, row in batch_df.iterrows():
        try:
            player_name = row['player_name'].strip()
            posteam = validate_team(row['posteam'])
            defteam = validate_team(row['defteam'])
            line = float(row['line'])
            is_home = int(row['is_home'])
            
            # Updated validation call with season/week
            validate_player(player_name, prop_name, player_data, season, prediction_week, min_games)
            
            result = predict_prop(prop_name, player_name, posteam, defteam, line, season, prediction_week, is_home)
            result['home_away'] = 'Home' if is_home else 'Away'
            results.append(result)
            print(f"✓ Predicted: {player_name} ({posteam} vs {defteam})")
            
        except Exception as e:
            print(f"✗ Error with {row.get('player_name', 'Unknown')}: {e}")
            continue
    
    return results

def main():
    season = 2025
    output_file = os.path.join('data', 'passing-prop-predictions-2025.xlsx')
    
    try:
        prediction_week = get_int_input("Enter the week to predict for (1-18): ", 1, 18)
        print(f"Predicting for Week {prediction_week} of season {season}...")
    except KeyboardInterrupt:
        print("\nExiting due to user interrupt.")
        return
    
    sheet_name = f'Week {prediction_week}'
    existing_players = get_existing_predictions(output_file, sheet_name)
    
    if existing_players:
        print(f"\nFound {len(existing_players)} existing predictions for Week {prediction_week}:")
        for player in sorted(existing_players):
            print(f"  - {player}")
        print("\nThese players will be skipped. New predictions will be added.")
    
    # Ask for batch or manual input
    print("\n" + "="*50)
    print("Choose input method:")
    print("1. Batch import from CSV")
    print("2. Manual entry")
    print("="*50)
    
    choice = input("Enter choice (1 or 2): ").strip()
    
    results = []
    prop_name = 'qb_passing_yards'
    
    try:
        player_data = pd.read_csv(os.path.join('data', f"{PROPS_CONFIG[prop_name]['target_stat']}_player_logs.csv"))
    except FileNotFoundError:
        print("Error: Player logs CSV not found. Run 'python -m src.train_pipeline' first.")
        return
    except Exception as e:
        print(f"Error loading CSV: {e}")
        return
    
    # Get min_games from config
    config = PROPS_CONFIG[prop_name]
    min_games = config.get('min_games', 10)
    
    if choice == '1':
        # Batch mode
        csv_path = input("Enter path to CSV file (e.g., data/week4_batch.csv): ").strip()
        try:
            batch_df = load_batch_csv(csv_path)
            print(f"\nLoaded {len(batch_df)} predictions from CSV")
            print("\nGenerating predictions...")
            results = batch_predict(batch_df, prop_name, season, prediction_week, player_data)
        except Exception as e:
            print(f"Error with batch import: {e}")
            return
    
    else:
        # Manual mode
        while True:
            print("\nEnter player details (or 'done' to finish):")
            try:
                player_name = input("Enter player name (e.g., P.Mahomes) or 'done': ").strip()
                if player_name.lower() == 'done':
                    break
                
                if player_name in existing_players:
                    print(f"Player {player_name} already exists in Week {prediction_week}. Skipping...")
                    continue
                
                # Updated validation call with season/week
                validate_player(player_name, prop_name, player_data, season, prediction_week, min_games)
                
                posteam = input("Enter player's team (e.g., KC): ").strip()
                posteam = validate_team(posteam)
                defteam = input("Enter opponent defense team (e.g., ATL): ").strip()
                defteam = validate_team(defteam)
                line = get_float_input(f"Enter Vegas {prop_name.replace('_', ' ')} line (e.g., 240.5): ")
                is_home = get_int_input("Is player home? (1 for yes, 0 for no): ", 0, 1)
                
                result = predict_prop(prop_name, player_name, posteam, defteam, line, season, prediction_week, is_home)
                result['home_away'] = 'Home' if is_home else 'Away'
                results.append(result)
                print(f"Added prediction for {player_name} ({posteam} vs {defteam}) - {'Home' if is_home else 'Away'}")
                
            except ValueError as e:
                print(f"Input error: {e}. Try again.")
            except KeyboardInterrupt:
                print("\nExiting input loop due to user interrupt.")
                break
            except Exception as e:
                print(f"Unexpected error: {e}. Skipping this prediction.")
    
    if results:
        print("\nGenerating predictions and saving to Excel...")
        df = pd.DataFrame(results)
        df['actual'] = None
        df['win_loss'] = None
        
        df['lean'] = df.apply(lambda row: 'OVER' if row['mean_prediction'] > row['line'] else 'UNDER', axis=1)
        
        df['prob_over_line'] = df['prob_over_line'].apply(lambda p: f"{p:.1%}")
        df['mean_prediction'] = df['mean_prediction'].round(1)
        df['p10'] = df['p10'].round(1)
        df['p50'] = df['p50'].round(1)
        df['p90'] = df['p90'].round(1)
        
        display_cols = ['player_name', 'posteam', 'defteam', 'home_away', 'mean_prediction', 'line', 'actual', 'prob_over_line', 'lean', 'win_loss', 'p10', 'p50', 'p90']
        print(f"\nWeek {prediction_week} New Predictions:")
        print(df[display_cols].to_string(index=False))
        
        save_to_excel(df[display_cols], prediction_week)
        
    else:
        if existing_players:
            print(f"No new predictions added. Week {prediction_week} still has {len(existing_players)} predictions.")
        else:
            print("No predictions generated.")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nProgram terminated by user.")
    except Exception as e:
        print(f"Fatal error: {e}")